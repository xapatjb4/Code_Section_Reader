import xlsxwriter as xl
import openpyxl as oxl
import os.path
from os import path
import glob
import sys

#List of strings to identify the code composition
def dict_Data(arg):
    switcher = {
        #Program Control Data
        0: 'su>', #SubRoutine
        1: 'fo>', #For loops
        2: 'wh>', #While loop
        3: 'if>', #if
        4: 'el>', #else
        5: 'ca>', #case
        6: 'br>', #break return go to
        #Variable Data
        7: 'va>', #Variable
        8: 'in>', #Integer
        9: 'fl>', #Float
        10: 'ch>', # String
        11: 'ar>', # Array (Not string)
        #Other Data
        12: 'co>', # Comment
        13: 'pr>' #programming language used 1 python 2 c/c++/c# 3 Java/javascript 4 swift 99 other

    }
    return switcher.get(arg, "nothing")

#Return instances of each word defined in word, in a line
def word_find(line,words):
    linetocheck = line.strip().split()
    #Making a set of words to be removed from the list to be counted
    toremove= set()
    for i,x in enumerate(linetocheck, start=0):
        if x not in words:
            toremove.add(x)
    #Removing the unmatch words from the list
    instances = [x for x in linetocheck if x not in toremove]
    return list(instances)

#Returns the lines which instance of words can be found
def word_line(file,words):
    with open(file) as f:
        for i,x in enumerate(f, start=1):
            common = word_find(x,words)
            if common:
                print(i, "".join(common))

#Returns a dictionnary containing the number of instances of words found
def word_count(file, words):
    #make a dict with words we're looking for
    var_tags = ['in>', 'fl>', 'ar>', 'ch>']
    counts = dict.fromkeys(words,0)
    with open(file) as f:
        for i,x in enumerate(f, start=0):
            #Check if line contains words being searched
            common = word_find(x,words)
            if common:
                #Increment number of instances for word found
                for word in common:
                    #possibly add check to increment var if in fl or ar
                    counts[word] += 1
                    if word in var_tags:
                        counts['va>'] += 1

    return counts

#Take in a list of variable names and return the average length
def list_avg(variables):
    #To see the list of vars uncomment this
    print(variables)
    if variables:
        return sum(len(word) for word in variables) / len(variables)
    return 0

#makes a list of whatever is tagged in va to return avg length
    #find va> then add var name to list
def var_avg(file):
    var_tags = ['in>', 'fl>', 'ar>', 'ch>', 'va>']
    var_names = []
    with open(file) as f:
        for i,x in enumerate(f, start=0):
            common = set(word_find(x,var_tags))
            if common:
                line = x.strip().split()
                #for every index containing va> add the following index to the list
                for x in range(0,len(line)):
                    if line[x] in var_tags:
                        #To print var line with var
                        print(i+1, line[x], line[x+1])
                        var_names.append(line[x+1])
                        x+=1
    return list_avg(var_names)

#Make a function to format data into excel sheet
def xl_format1(workbookName):
    #Check if file exists
    #if path.exists(outputName):
    #    return -1
    #Format should be as follows
    #worksheet1
    #col 0 subject, 1 su, 2 for, 3 while 4 if 5 else 6 case 7 break
    #worksheet2
    #col 0 subject 1 #var 2 #int 3 #float 4 #string 5 #array 6 avg varname 7 comment
    if os.path.exists(workbookName):
        print("File with that name already exists")
        return -1
    wb = oxl.Workbook()
    xl_PC_page(wb)
    xl_Var_page(wb)
    xl_Oth_page(wb)
    #Remove Default sheet
    std = wb['Sheet']
    wb.remove(std)
    wb.save(filename = workbookName)

    return 0

def xl_format2(outputName):
    return 0

def xl_PC_page(workbook):
    ws = workbook.create_sheet('Program Control')
    ws['A1'] = 'Subject'
    ws['B1'] = 'Subroutine'
    ws['C1'] = 'For'
    ws['D1'] = 'While'
    ws['E1'] = 'If'
    ws['F1'] = 'Else'
    ws['G1'] = 'Case'
    ws['H1'] = 'GoTo'

    #workbook.save(workbook)
    return 0

def xl_Var_page(workbook):
    ws = workbook.create_sheet('Variables')
    ws['A1'] = 'Subject'
    ws['B1'] = '# Vars'
    ws['C1'] = 'Ave Len'
    #workbook.save(workbook)
    return 0
    #old Format
    '''Vworksheet.write('A1', 'Subject')
    Vworksheet.write('B1', '# Vars')
    Vworksheet.write('C1', '# Ints')
    Vworksheet.write('D1', '# Floats')
    Vworksheet.write('E1', '# Chars')
    Vworksheet.write('F1', '# Arrays')
    Vworksheet.write('G1', 'Ave Len')
    Vworksheet.write('H1', 'Comments')
    workbook.close()
    wb = oxl.load_workbook(outputName)'''

def xl_Oth_page(workbook):
    ws = workbook.create_sheet('Other Data')
    ws['A1'] = 'Subject'
    ws['B1'] = 'Comments'
    ws['C1'] = 'Prog Lang'
    #workbook.save(workbook)
    return 0

#Fill the excel sheet with occurences found
def xl_fill(outfile, subject, occurences, var_avg):
    #take care of program control variable
    #Modify the xcel sheet to contain the respective instance
    wb = oxl.load_workbook(outfile)
    ws = wb['Program Control']
    ws.cell(subject+1, 1).value = subject

    x = 0
    for x in range(0,7):
        #Modify row {sujbect}  with war data 0 through 7
        value = occurences.get(dict_Data(x))
        ws.cell(subject+1, x+2).value = value

    ws = wb['Variables']
    ws.cell(subject+1, 1).value = subject
    #get num var, avg len
    ws.cell(subject+1, 2).value = occurences.get('va>')
    ws.cell(subject+1, 3).value = var_avg


    ws = wb['Other Data']
    ws.cell(subject+1, 1).value = subject
    ws.cell(subject+1, 2).value = occurences.get('co>')
    #checking language
    progLan = occurences.get('pr>')
    if progLan > 4:
        proLan = 99
    ws.cell(subject+1, 3).value = progLan

    wb.save(outfile)
    return 0;


if __name__ == '__main__':
    outputFile = sys.argv[2]
    xl_format1(outputFile)
    inputDir = sys.argv[1] + "/*.txt"
    #print(inputDir, outputFile)
    words = []
    for x in range(14):
        words.append(dict_Data(x))
    files = glob.glob(inputDir)
    #Split the string to remove
    for file in files:
        split_file = file.split('/')# take the / from path
        subject = int(split_file[1][:-4])# Taking out the .txt
        print(subject)
        xl_fill(outputFile, subject, word_count(file, words),var_avg(file))
        #word_line(file,words)
        print('-----------------------------------------------------------')
    #word_line(file, words)
    #print(word_count(file, words))
